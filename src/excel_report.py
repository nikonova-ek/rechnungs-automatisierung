import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Daten aus SQLite laden
conn = sqlite3.connect("output/rechnungen.db")
df = pd.read_sql("SELECT * FROM rechnungen", conn)
df["brutto"] = pd.to_numeric(df["brutto"], errors="coerce")
df["netto"] = pd.to_numeric(df["netto"], errors="coerce")
df["mwst"] = pd.to_numeric(df["mwst"], errors="coerce")
conn.close()

print(df)
print(f"\nAnzahl Datensätze: {len(df)}")

# Excel-Datei erstellen
wb = Workbook()

# Blatt 1: Daten
ws1 = wb.active
ws1.title = "Daten"

# Spaltenüberschriften
headers = ["Rechnungsnr.", "Datum", "Lieferant", "Netto", "MwSt", "Brutto"]
for col, header in enumerate(headers, 1):
    ws1.cell(row=1, column=col, value=header)

# Daten einfügen
for row_idx, row in df.iterrows():
    ws1.cell(row=row_idx+2, column=1, value=row["rechnungsnr"])
    ws1.cell(row=row_idx+2, column=2, value=str(row["datum"])[:10])
    ws1.cell(row=row_idx+2, column=3, value=row["lieferant"])
    ws1.cell(row=row_idx+2, column=4, value=row["netto"])
    ws1.cell(row=row_idx+2, column=5, value=row["mwst"])
    ws1.cell(row=row_idx+2, column=6, value=row["brutto"])

# Fettschrift für Überschriften
for col in range(1, 7):
    ws1.cell(row=1, column=col).font = Font(bold=True)

# Spaltenbreite anpassen
column_widths = [20, 15, 35, 12, 12, 12]
for col, width in enumerate(column_widths, 1):
    ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = width
# Blatt 2: Auswertung
ws2 = wb.create_sheet("Auswertung")

# Ausgaben pro Lieferant
ws2.cell(row=1, column=1, value="Ausgaben pro Lieferant")
ws2.cell(row=1, column=1).font = Font(bold=True)

ws2.cell(row=2, column=1, value="Lieferant")
ws2.cell(row=2, column=2, value="Gesamt (Brutto)")
ws2.cell(row=2, column=1).font = Font(bold=True)
ws2.cell(row=2, column=2).font = Font(bold=True)

lieferant_summen = df.groupby("lieferant")["brutto"].sum().sort_values(ascending=False)
for row_idx, (lieferant, summe) in enumerate(lieferant_summen.items(), 3):
    ws2.cell(row=row_idx, column=1, value=lieferant)
    ws2.cell(row=row_idx, column=2, value=round(summe, 2))

# Ausgaben pro Monat
start_row = len(lieferant_summen) + 5
ws2.cell(row=start_row, column=1, value="Ausgaben pro Monat")
ws2.cell(row=start_row, column=1).font = Font(bold=True)

ws2.cell(row=start_row+1, column=1, value="Monat")
ws2.cell(row=start_row+1, column=2, value="Gesamt (Brutto)")
ws2.cell(row=start_row+1, column=1).font = Font(bold=True)
ws2.cell(row=start_row+1, column=2).font = Font(bold=True)

df["monat"] = df["datum"].str[:7]
monat_summen = df.groupby("monat")["brutto"].sum()
for row_idx, (monat, summe) in enumerate(monat_summen.items(), start_row+2):
    ws2.cell(row=row_idx, column=1, value=monat)
    ws2.cell(row=row_idx, column=2, value=round(summe, 2))

ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 20
# Bedingte Formatierung — teuerster Lieferant rot markieren
red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
max_row = lieferant_summen.index.get_loc(lieferant_summen.idxmax()) + 3
ws2.cell(row=max_row, column=1).fill = red_fill
ws2.cell(row=max_row, column=2).fill = red_fill
# Datei speichern
wb.save("output/rechnungen.xlsx")
print("Excel-Datei gespeichert!")